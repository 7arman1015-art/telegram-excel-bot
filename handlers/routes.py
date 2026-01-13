from aiogram import Router
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

router = Router()


@router.message(Command("start"))
async def start(message: Message):
    await message.answer(
        "Привет 👋\n"
        "Я делаю Excel файл из текста.\n"
        "Напиши /info"
    )


@router.message(Command("info"))
async def info(message: Message):
    await message.answer(
        "Отправь текст в формате:\n\n"
        "человек_паук\n"
        "железный_человек\n\n"
        "Я верну Excel файл, где:\n"
        "человек → A1\n"
        "паук → B1\n"
        "железный → A2\n"
        "человек → B2"
    )


@router.message()
async def text_handler(message: Message):
    text = message.text

    # 1️⃣ Создаём Excel
    wb = Workbook()
    ws = wb.active

    # 2️⃣ Стили
    purple_fill = PatternFill(
        start_color="800080",
        end_color="800080",
        fill_type="solid"
    )

    wrap_alignment = Alignment(wrap_text=True)

    # 3️⃣ Обрабатываем текст
    lines = text.splitlines()
    row = 1

    for line in lines:
        if "_" not in line:
            continue

        left, right = line.split("_", 1)

        ws[f"A{row}"] = left
        ws[f"B{row}"] = right

        ws[f"A{row}"].alignment = wrap_alignment
        ws[f"B{row}"].alignment = wrap_alignment

        row += 1

    # 4️⃣ Фиолетовая первая строка
    ws["A1"].fill = purple_fill
    ws["B1"].fill = purple_fill

    # 5️⃣ Ширина колонок
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15

    # 6️⃣ Сохраняем файл
    filename = "result.xlsx"
    wb.save(filename)

    # 7️⃣ Отправляем файл
    file = FSInputFile(filename)
    await message.answer_document(
        document=file,
        caption="Готово ✅"
    )
